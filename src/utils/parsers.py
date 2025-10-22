# Copyright (C) 2025-present by FringeLabs@Github, < https://github.com/FringeLabs >.
#
# This file is part of < https://github.com/FringeLabs/FLASK-PRP > project,
# and is released under the "GNU v3.0 License Agreement".
# Please see < https://github.com/FringeLabs/FLASK-PRP/blob/main/LICENSE >
#
# All rights reserved.

import re
import yaml
from .basic_utils import *
import os
import json
from openpyxl import Workbook
from collections import Counter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


class TextParser:

    def __init__(
        self, text: str, save_folder: str = r"output/", college_code: str = "NSS"
    ):
        self.text = text
        self._init = None
        self.yaml_path = "src/utils/file_utils/regex.yaml"
        self.regex_data = {}
        self.supported_STREAMS = load_subjects("src/utils/file_utils/supported_streams.yaml")
        self.college_code = college_code
        self.d_time = datetime.now()
        self.GRADE_ORDER = ["S", "A+", "A", "B+", "B", "C+", "C", "D", "P", "Absent", "F"]
        self.save_folder = save_folder
        self.subj_CODE = {}
        self.exam_centre = None
        self.exam_name = None
        self.init

    @property
    def init(self):
        if not self._init:
            with open(self.yaml_path, encoding="utf-8") as fr:
                self.regex_data = yaml.safe_load(fr)
            self.format_match_check()
            self.subj_CODE = self.extract_course_dict()
        self._init = True
        return self._init
    
    def text_matcher_year_wise(self):
        n_data = self.supported_STREAMS
        result_year_wise = {}
        result_num_data_year_wise = {}

        for stream_code, subject_name in n_data.items():
            pattern = re.compile(
                rf"({self.college_code})(\d{{2}}){stream_code}(\d+)\s+((?:\w+\(([A-Z][+A-Z]*|Absent)\),?\s*)+)",
                re.MULTILINE,
            )
            course_grade_pattern = re.compile(self.regex_data["course_code_pattern"])

            for match in pattern.finditer(self.text):
                year_code = match.group(2) 
                student_suffix = match.group(3)
                subjects_str = match.group(4)
                student_id = f"{self.college_code}{year_code}{stream_code}{student_suffix}"

                subjects = dict(course_grade_pattern.findall(subjects_str))
                nstream_code = self.supported_STREAMS[stream_code]
                result_year_wise.setdefault(year_code, {}).setdefault(nstream_code, {})
                result_num_data_year_wise.setdefault(year_code, {}).setdefault(nstream_code, {
                    "appeared": 0,
                    "passed": 0,
                    "subjects": {},
                    "grade_count": {},
                })

                result_num_data_year_wise[year_code][nstream_code]["appeared"] += 1
                if all(grade not in ["F", "Absent"] for grade in subjects.values()):
                    result_num_data_year_wise[year_code][nstream_code]["passed"] += 1

                for scode, grade in subjects.items():
                    result_num_data_year_wise[year_code][nstream_code]["grade_count"][grade] = (
                        result_num_data_year_wise[year_code][nstream_code]["grade_count"].get(grade, 0) + 1
                    )

                    if scode not in result_num_data_year_wise[year_code][nstream_code]["subjects"]:
                        result_num_data_year_wise[year_code][nstream_code]["subjects"][scode] = {
                            "appeared": 0,
                            "passed": 0,
                        }

                    result_num_data_year_wise[year_code][nstream_code]["subjects"][scode]["appeared"] += 1
                    if grade not in ["F", "AB"]:
                        result_num_data_year_wise[year_code][nstream_code]["subjects"][scode]["passed"] += 1

                    grade_key = f"no_of_{grade}_grades"
                    result_num_data_year_wise[year_code][nstream_code]["subjects"][scode][grade_key] = (
                        result_num_data_year_wise[year_code][nstream_code]["subjects"][scode].get(grade_key, 0)
                        + 1
                    )

                subjects["Arrear Subjects Count"] = sum(
                    1 for grade in subjects.values() if grade in ["F", "Absent"]
                )
                subjects["Arrear Subject Codes"] = ", ".join(
                    [code for code, grade in subjects.items() if grade in ["Absent", "F"]]
                )

                result_year_wise[year_code][nstream_code][student_id] = subjects

        return result_year_wise, result_num_data_year_wise

    def extract_course_dict(self):
        course_dict = {}
        lines = self.text.splitlines()
        course_pattern = re.compile(r'^([A-Z]{2,4}[A-Z]*\d{3})\s+(.*)')
        current_code = None
        current_name = []
        capture = False
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if "Course Code Course" in line:
                capture = True
                continue
            if line.startswith("Register No"):
                capture = False
                continue
            if capture:
                match = course_pattern.match(line)
                if match:
                    if current_code:
                        course_dict[current_code] = ' '.join(current_name).strip()
                    current_code = match.group(1)
                    current_name = [match.group(2)]
                else:
                    if current_code:
                        current_name.append(line)
        if current_code:
            course_dict[current_code] = ' '.join(current_name).strip()

        return course_dict


    def format_match_check(self) -> None:
        c = re.search(self.regex_data["exam_centre_pattern"], self.text)
        e = re.search(self.regex_data["btech_exam_pattern"], self.text)
        if not c:
            raise ValueError("Exam Centre pattern not found")
        if not e:
            raise ValueError("B.Tech Exam pattern not found")
        if not self.college_code:
            matches = re.findall(self.regex_data['college_code_pattern'], self.text)
            if not matches:
                raise ValueError("College code could not be predicted from text")
            most_common = Counter(matches).most_common(1)[0][0]
            self.college_code = most_common
        self.exam_centre, self.exam_name = c.group(1), e.group(1)

    def parse_data_as_json_file(self):
        result = self.text_matcher_year_wise()
        save_folder = self.__make_dir_if_not_exists()
        json_file_path = os.path.join(save_folder, "results.json")
        with open(json_file_path, "w", encoding="utf-8") as jfile:
            json.dump(result, jfile, indent=4)
        return json_file_path

    def __make_dir_if_not_exists(self):
        if not os.path.exists(self.save_folder):
            os.makedirs(self.save_folder)
        return self.save_folder

    def extract_data_and_return_as_csv(self):
        result_year_wise, rnd_year_wise = self.text_matcher_year_wise()
        save_folder = self.__make_dir_if_not_exists()
        exam_centre, exam_name = self.exam_centre, self.exam_name
        stream_subjects = self.supported_STREAMS
        for stream_code in stream_subjects.keys():
            xstream_code = self.supported_STREAMS[stream_code]
            xlsx_file_path = os.path.join(save_folder, f"{xstream_code.replace(" ", "_")}_results.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append([f"Exam Centre: {exam_centre}"])
            ws.append([f"Exam Name: {exam_name}"])
            for row in [ws.max_row - 1, ws.max_row]:
                cell = ws.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            ws.append([])
            ws.append([f"Stream: {xstream_code} (Generated on: {self.d_time.strftime('%Y-%m-%d %H:%M:%S')})"])
            cell = ws.cell(row=ws.max_row, column=1)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.append([])
            mixed_students = {}
            mixed_rnd = {"appeared":0, "passed":0, "grade_count":{}, "subjects":{}}

            for year_code, streams in result_year_wise.items():
                year_students = streams.get(xstream_code, {})
                mixed_students.update(year_students)
                year_rnd = rnd_year_wise.get(year_code, {}).get(xstream_code, {})
                mixed_rnd["appeared"] += year_rnd.get("appeared",0)
                mixed_rnd["passed"] += year_rnd.get("passed",0)
                for g, count in year_rnd.get("grade_count", {}).items():
                    mixed_rnd["grade_count"][g] = mixed_rnd["grade_count"].get(g,0) + count
                for subject_code, stats in year_rnd.get("subjects", {}).items():
                    if subject_code not in mixed_rnd["subjects"]:
                        mixed_rnd["subjects"][subject_code] = stats.copy()
                    else:
                        for key, value in stats.items():
                            mixed_rnd["subjects"][subject_code][key] = mixed_rnd["subjects"][subject_code].get(key,0) + value
            sections = [("MIXED (All Years)", mixed_students, mixed_rnd)]
            for year_code, streams in result_year_wise.items():
                sections.append((f"YEAR {year_code}", streams.get(xstream_code, {}), rnd_year_wise.get(year_code, {}).get(xstream_code, {})))
            for section_name, students, rnd in sections:
                if not students:
                    continue
                ws.append([section_name])
                cell = ws.cell(row=ws.max_row, column=1)
                cell.font = Font(bold=True, color="000000")
                cell.alignment = Alignment(horizontal="left")
                ws.append([])
                fieldnames = ["Student ID"] + list(next(iter(students.values()), {}).keys())
                ws.append(fieldnames)
                for col in range(1, len(fieldnames) + 1):
                    cell = ws.cell(row=ws.max_row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="e60000", end_color="e60000", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for student_id, subjects in students.items():
                    row = [student_id] + [subjects.get(sub, "") for sub in fieldnames[1:]]
                    ws.append(row)
                    for col in range(2, len(row) + 1):
                        ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="center")
                total_appeared = rnd.get("appeared", sum(1 for s in students.values()))
                total_passed = rnd.get("passed", sum(1 for s in students.values() if all(g not in ["F", "Absent"] for g in s.values())))
                pass_percent = (total_passed / total_appeared * 100) if total_appeared > 0 else 0
                summary_rows = [
                    ["=== Section Summary ==="],
                    [f"Total Students Appeared: {total_appeared}"],
                    [f"Total Students Passed: {total_passed}"],
                    [f"Pass Percentage: {pass_percent:.2f}%"]
                ]
                for row_data in summary_rows:
                    ws.append(row_data)
                    for col in range(1, len(row_data) + 1):
                        cell = ws.cell(row=ws.max_row, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")
                ws.append([])
                grade_count = rnd.get("grade_count", {})
                all_grades = [g for g in self.GRADE_ORDER if g in grade_count]
                ws.append(["GRADE WISE COUNT"])
                ws.append(["Grade"] + all_grades)
                ws.append(["Count"] + [grade_count.get(g, 0) for g in all_grades])
                for col in range(1, len(all_grades) + 2):
                    cell = ws.cell(row=ws.max_row - 1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="e60000", end_color="e60000", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                for row in range(ws.max_row - 1, ws.max_row + 1):
                    for col in range(2, len(all_grades) + 2):
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                ws.append([])
                ws.append(["SUBJECT WISE PERFORMANCE STATISTICS"])
                subject_table_headers = ["Subject Code", "Appeared", "Passed", "Pass %"] + all_grades
                ws.append(subject_table_headers)
                for col in range(1, len(subject_table_headers) + 1):
                    cell = ws.cell(row=ws.max_row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="e60000", end_color="e60000", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                subjects_stats = rnd.get("subjects", {})
                for subject_code, stats in subjects_stats.items():
                    pass_percentage = (stats["passed"] / stats["appeared"] * 100) if stats["appeared"] > 0 else 0
                    row = [subject_code, stats["appeared"], stats["passed"], f"{pass_percentage:.2f}"]
                    for grade in all_grades:
                        grade_key = f"no_of_{grade}_grades"
                        row.append(stats.get(grade_key, 0))
                    ws.append(row)
                    for col in range(2, len(row) + 1):
                        ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="center")
                ws.append([])
            for col in range(1, ws.max_column + 1):
                max_length = 0
                for cell in ws[get_column_letter(col)]:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[get_column_letter(col)].width = max_length + 2
            wb.save(xlsx_file_path)
        return save_folder